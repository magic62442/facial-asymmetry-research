#!/usr/bin/env python
# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

csv_files = {
    'bijian_GT': 'bijian/ground_truth/ground_truth_summary_statistics.csv',
    'bijian_ICP': 'bijian/icp_result/icp_summary_statistics.csv',
    'bijian_MeshMonk': 'bijian/meshmonk/meshmonk_summary_statistics.csv',
    'kedian_GT': 'kedian/ground_truth/ground_truth_summary_statistics.csv',
    'kedian_ICP': 'kedian/icp_result/icp_summary_statistics.csv',
    'kedian_MeshMonk': 'kedian/meshmonk/meshmonk_summary_statistics.csv',
    'ppdh_MeshMonk': 'analysis_output_ppdh_mapped_templates/summary_statistics.csv',
    'headspace_MeshMonk': 'analysis_output_headspace_mapped_templates/summary_statistics.csv',
}

wb = Workbook()
ws_all = wb.active
ws_all.title = 'All_Results'

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='4472C4')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

all_data = []
for sheet_name, csv_path in csv_files.items():
    if os.path.exists(csv_path):
        df = pd.read_csv(csv_path)
        parts = sheet_name.split('_')
        df.insert(0, 'Method', parts[-1])
        df.insert(0, 'Dataset', parts[0])
        all_data.append(df)

        ws = wb.create_sheet(title=sheet_name)
        df_sheet = pd.read_csv(csv_path)
        for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

if all_data:
    combined_df = pd.concat(all_data, ignore_index=True)
    for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_all.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
    for col in ws_all.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws_all.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)

output_path = 'summary_statistics_all.xlsx'
wb.save(output_path)
print(f"Excel文件已保存: {output_path}")
print(f"包含 {len(wb.sheetnames)} 个sheet: {', '.join(wb.sheetnames)}")
